"""
Conference API Server

Provides endpoints for:
- Conference list and metadata
- Conference scholars list
- Scholar details
- Avatar proxy with caching
- Label filtering
"""

import hashlib
import json
import logging
from datetime import datetime
from functools import lru_cache
from io import BytesIO
from pathlib import Path
from typing import Optional
from urllib.parse import urlparse

import httpx
from fastapi import FastAPI, HTTPException, Query, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

from config import settings

# Configure logging
logging.basicConfig(
    level=getattr(logging, settings.log_level.upper()),
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

app = FastAPI(title="Conference API", version="0.1.0")

# CORS configuration
app.add_middleware(
    CORSMiddleware,
    allow_origins=settings.cors_origins_list,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# HTTP client for fetching remote avatars
http_client = httpx.AsyncClient(timeout=settings.http_timeout, follow_redirects=True)

# Cache timeout tracking
_cache_timestamp: dict[str, float] = {}


def get_cache_key(func_name: str, *args) -> str:
    """Generate a cache key for function calls."""
    return f"{func_name}:{':'.join(str(a) for a in args)}"


def is_cache_valid(key: str) -> bool:
    """Check if cache entry is still valid."""
    if key not in _cache_timestamp:
        return False
    return (datetime.now().timestamp() - _cache_timestamp[key]) < settings.cache_ttl_seconds


def update_cache_timestamp(key: str):
    """Update cache timestamp."""
    _cache_timestamp[key] = datetime.now().timestamp()


@lru_cache(maxsize=128)
def _load_json_file(file_path: str) -> dict:
    """Load and cache a JSON file."""
    with open(file_path, "r", encoding="utf-8") as f:
        return json.load(f)


def load_json_file(file_path: str) -> dict:
    """Load a JSON file with cache validation."""
    cache_key = get_cache_key("load_json", file_path)
    if not is_cache_valid(cache_key):
        _load_json_file.cache_clear()
    update_cache_timestamp(cache_key)
    return _load_json_file(file_path)


def get_aminer_avatar_url(aminer_id: str) -> str:
    """Generate AMiner avatar URL."""
    return f"https://static.aminer.cn/upload/avatar/{aminer_id}.jpg"


def get_avatar_cache_path(url: str) -> Path:
    """Generate cache file path for a given URL."""
    url_hash = hashlib.md5(url.encode()).hexdigest()
    # Extract extension from URL, default to .jpg
    parsed = urlparse(url)
    ext = Path(parsed.path).suffix or ".jpg"
    return settings.avatar_cache_dir / f"{url_hash}{ext}"


def get_avatar_fail_marker_path(url: str) -> Path:
    """Generate fail marker file path for a given URL."""
    url_hash = hashlib.md5(url.encode()).hexdigest()
    return settings.avatar_cache_dir / f"{url_hash}.failed"


def is_avatar_fetch_failed(url: str) -> bool:
    """Check if avatar fetch has failed recently (within TTL)."""
    fail_marker = get_avatar_fail_marker_path(url)
    if not fail_marker.exists():
        return False
    # Check if marker is expired
    marker_age = datetime.now().timestamp() - fail_marker.stat().st_mtime
    if marker_age > settings.avatar_fail_cache_ttl:
        fail_marker.unlink()  # Remove expired marker
        return False
    return True


def mark_avatar_fetch_failed(url: str):
    """Mark avatar fetch as failed."""
    fail_marker = get_avatar_fail_marker_path(url)
    fail_marker.touch()


def get_local_avatar_path(aminer_id: str) -> Optional[Path]:
    """
    Check if avatar exists locally in data/aminer/avatars and is not a default avatar.
    Returns the path to the avatar file if found.
    """
    # Check for .default marker - if exists, this scholar has default avatar
    default_marker = settings.aminer_avatars_dir / f"{aminer_id}.default"
    if default_marker.exists():
        return None

    # Check for avatar files (try all extensions)
    for ext in ['.jpg', '.jpeg', '.png']:
        avatar_path = settings.aminer_avatars_dir / f"{aminer_id}{ext}"
        if avatar_path.exists():
            return avatar_path

    return None


def get_scholar_photo_url(aminer_id: Optional[str]) -> Optional[str]:
    """
    Get the original photo URL for a scholar.
    Priority: enriched photo_url > AMiner avatar > None
    """
    if not aminer_id:
        return None

    # Check enriched data for photo_url first
    enriched_path = settings.enriched_scholars_dir / f"{aminer_id}.json"
    if enriched_path.exists():
        try:
            enriched_data = load_json_file(str(enriched_path))
            if enriched_data.get("photo_url"):
                return enriched_data["photo_url"]
        except Exception:
            pass

    # Fall back to AMiner avatar
    return get_aminer_avatar_url(aminer_id)


def get_scholar_photo(aminer_id: Optional[str]) -> Optional[str]:
    """
    Get scholar photo URL (proxied through our API for caching).
    Returns the proxy URL that will cache and serve the avatar.

    Priority:
    1. Local avatar in data/aminer/avatars (if exists and not default)
    2. Our own cached/proxied avatar
    """
    if not aminer_id:
        return None

    # Check if we have a local avatar
    local_avatar_path = get_local_avatar_path(aminer_id)
    if local_avatar_path:
        # Return URL to serve the local avatar
        return f"/api/avatar/local/{aminer_id}"

    # Fall back to proxy URL - frontend will request this, and we'll cache the avatar
    return f"/api/avatar/{aminer_id}"


# ============== Pydantic Models ==============

class ConferenceLocation(BaseModel):
    venue: Optional[str] = None
    city: Optional[str] = None
    country: Optional[str] = None


class ConferenceDates(BaseModel):
    start: Optional[str] = None
    end: Optional[str] = None


class ConferenceUrl(BaseModel):
    url: str
    name: Optional[str] = None
    description: Optional[str] = None


class ConferenceMeta(BaseModel):
    id: str  # directory name as ID
    name: str
    shortName: Optional[str] = None
    edition: Optional[int] = None
    year: Optional[int] = None
    dates: Optional[ConferenceDates] = None
    location: Optional[ConferenceLocation] = None
    description: Optional[str] = None
    logo_url: Optional[str] = None
    urls: Optional[list[ConferenceUrl]] = None
    timezone: Optional[str] = None
    tags: Optional[list[str]] = None


class ScholarBasic(BaseModel):
    name: str
    name_zh: Optional[str] = None
    affiliation: Optional[str] = None
    roles: list[str] = []
    aminer_id: Optional[str] = None
    photo_url: Optional[str] = None
    description: Optional[str] = None


class AminerValidation(BaseModel):
    status: Optional[str] = None
    is_same_person: Optional[bool] = None
    reason: Optional[str] = None


# Labels related models
class LabelResult(BaseModel):
    name: str
    value: Optional[bool] = None
    confidence: Optional[str] = None
    reason: Optional[str] = None


class ScholarLabels(BaseModel):
    last_updated: Optional[str] = None
    results: list[LabelResult] = []


class LabelDefinition(BaseModel):
    name: str
    description: str


class LabelsConfig(BaseModel):
    version: str
    labels: list[LabelDefinition]


class ConferencePaperAuthor(BaseModel):
    name: str
    aminer_id: Optional[str] = None
    in_conference: bool = False


class ConferencePaper(BaseModel):
    paper_id: str
    title: str
    track: Optional[str] = None
    session: Optional[str] = None
    room: Optional[str] = None
    date: Optional[str] = None
    presentation_type: Optional[str] = None
    authors: list[ConferencePaperAuthor] = []  # All authors in original order (including current scholar)
    coauthors: list[ConferencePaperAuthor] = []  # Co-authors only (excluding current scholar)
    author_position: Optional[int] = None  # Position of current scholar in author list (1-indexed)
    abstract: Optional[str] = None


class AcademicIndices(BaseModel):
    hindex: Optional[float] = None
    gindex: Optional[float] = None
    citations: Optional[int] = None
    pubs: Optional[int] = None
    activity: Optional[float] = None
    diversity: Optional[float] = None
    sociability: Optional[float] = None


class ScholarDetail(BaseModel):
    # Basic info from scholars.json
    name: str
    name_zh: Optional[str] = None
    aliases: Optional[list[str]] = None
    affiliation: Optional[str] = None
    roles: list[str] = []
    description: Optional[str] = None
    sources: Optional[list[str]] = None
    source_urls: Optional[list[str]] = None
    aminer_id: Optional[str] = None
    aminer_validation: Optional[AminerValidation] = None
    photo_url: Optional[str] = None

    # From AMiner data
    bio: Optional[str] = None
    education: Optional[str] = None
    position: Optional[str] = None
    organizations: Optional[list[str]] = None
    honors: Optional[list[dict]] = None
    research_interests: Optional[list[dict]] = None

    # From enriched data
    homepage: Optional[str] = None
    google_scholar: Optional[str] = None
    dblp: Optional[str] = None
    linkedin: Optional[str] = None
    twitter: Optional[str] = None
    email: Optional[str] = None
    orcid: Optional[str] = None
    semantic_scholar: Optional[str] = None
    additional_info: Optional[str] = None

    # Labels
    labels: Optional[ScholarLabels] = None

    # Academic indices from enriched data
    indices: Optional[AcademicIndices] = None

    # Conference papers
    conference_papers: Optional[list[ConferencePaper]] = None


class HealthResponse(BaseModel):
    status: str
    timestamp: str


# ============== API Endpoints ==============

@app.get("/health", response_model=HealthResponse)
def health_check():
    """Health check endpoint."""
    return HealthResponse(
        status="healthy",
        timestamp=datetime.now().isoformat()
    )


@app.get("/api/conferences", response_model=list[ConferenceMeta])
def get_conferences():
    """
    Get all conferences.
    Scans data directory for subdirectories containing meta.json.
    """
    conferences = []

    if not settings.data_dir.exists():
        raise HTTPException(status_code=500, detail=f"Data directory not found: {settings.data_dir}")

    for item in settings.data_dir.iterdir():
        if item.is_dir():
            meta_path = item / "meta.json"
            if meta_path.exists():
                try:
                    meta_data = load_json_file(str(meta_path))
                    meta_data["id"] = item.name
                    conferences.append(ConferenceMeta(**meta_data))
                except Exception as e:
                    print(f"Error loading {meta_path}: {e}")
                    continue

    return conferences


@app.get("/api/conferences/{conference_id}/scholars", response_model=list[ScholarBasic])
def get_conference_scholars(conference_id: str):
    """
    Get scholars for a specific conference.
    Returns list of scholars with basic info and photo URLs.
    """
    conference_dir = settings.data_dir / conference_id
    scholars_path = conference_dir / "scholars.json"

    if not conference_dir.exists():
        raise HTTPException(status_code=404, detail=f"Conference not found: {conference_id}")

    if not scholars_path.exists():
        raise HTTPException(status_code=404, detail=f"Scholars data not found for conference: {conference_id}")

    try:
        scholars_data = load_json_file(str(scholars_path))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error loading scholars data: {e}")

    # Handle both formats: { "talents": [...] } or { "metadata": {...}, "talents": [...] }
    talents = scholars_data.get("talents", [])

    scholars = []
    for talent in talents:
        aminer_id = talent.get("aminer_id")
        photo_url = get_scholar_photo(aminer_id)

        # Try to get Chinese name from AMiner data
        name_zh = None
        if aminer_id:
            aminer_path = settings.aminer_scholars_dir / f"{aminer_id}.json"
            if aminer_path.exists():
                try:
                    aminer_data = load_json_file(str(aminer_path))
                    name_zh = aminer_data.get("detail", {}).get("name_zh")
                except Exception:
                    pass

        scholars.append(ScholarBasic(
            name=talent.get("name", "Unknown"),
            name_zh=name_zh,
            affiliation=talent.get("affiliation"),
            roles=talent.get("roles", []),
            aminer_id=aminer_id,
            photo_url=photo_url,
            description=talent.get("description"),
        ))

    return scholars


@app.get("/api/conferences/{conference_id}/data/scholars")
def get_conference_scholars_data(conference_id: str):
    """
    Get raw scholars data for a specific conference.
    Returns data from scholars.json file.
    """
    conference_dir = settings.data_dir / conference_id
    scholars_path = conference_dir / "scholars.json"

    if not conference_dir.exists():
        raise HTTPException(status_code=404, detail=f"Conference not found: {conference_id}")

    if not scholars_path.exists():
        raise HTTPException(status_code=404, detail=f"Scholars data not found for conference: {conference_id}")

    try:
        scholars_data = load_json_file(str(scholars_path))
        return scholars_data
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error loading scholars data: {e}")


@app.get("/api/conferences/{conference_id}/authors")
def get_conference_authors(conference_id: str):
    """
    Get authors for a specific conference (paper authors with metrics).
    Returns data from authors.json file with photo_url added for each author.
    """
    conference_dir = settings.data_dir / conference_id
    authors_path = conference_dir / "authors.json"

    if not conference_dir.exists():
        raise HTTPException(status_code=404, detail=f"Conference not found: {conference_id}")

    if not authors_path.exists():
        raise HTTPException(status_code=404, detail=f"Authors data not found for conference: {conference_id}")

    try:
        authors_data = load_json_file(str(authors_path))

        # Add photo_url and name_zh for each author (prioritizing local avatars)
        authors = authors_data.get("authors", [])
        for author in authors:
            aminer_id = author.get("aminer_id")
            author["photo_url"] = get_scholar_photo(aminer_id)

            # Try to get Chinese name from AMiner data
            if aminer_id:
                aminer_path = settings.aminer_scholars_dir / f"{aminer_id}.json"
                if aminer_path.exists():
                    try:
                        aminer_data = load_json_file(str(aminer_path))
                        name_zh = aminer_data.get("detail", {}).get("name_zh")
                        if name_zh:
                            author["name_zh"] = name_zh
                    except Exception:
                        pass

        return authors_data
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error loading authors data: {e}")


@app.get("/api/conferences/{conference_id}/scholars/search", response_model=list[ScholarDetail])
def search_scholars(
    conference_id: str,
    name: Optional[str] = Query(None, description="Scholar name to search"),
    aminer_id: Optional[str] = Query(None, description="AMiner ID to search"),
):
    """
    Search for scholars by name and/or aminer_id (OR logic).
    Returns detailed scholar information.
    Searches in both scholars.json (conference organizers) and authors.json (paper authors).
    """
    if not name and not aminer_id:
        raise HTTPException(status_code=400, detail="At least one of 'name' or 'aminer_id' must be provided")

    conference_dir = settings.data_dir / conference_id

    if not conference_dir.exists():
        raise HTTPException(status_code=404, detail=f"Conference not found: {conference_id}")

    # Search in scholars.json first (conference organizers)
    matching_talents = []
    scholars_path = conference_dir / "scholars.json"
    if scholars_path.exists():
        try:
            scholars_data = load_json_file(str(scholars_path))
            talents = scholars_data.get("talents", [])

            for talent in talents:
                match = False
                if name and talent.get("name", "").lower() == name.lower():
                    match = True
                if aminer_id and talent.get("aminer_id") == aminer_id:
                    match = True
                if match:
                    matching_talents.append(talent)
        except Exception as e:
            print(f"Error loading scholars data: {e}")

    # If not found in scholars.json, search in authors.json (paper authors)
    if not matching_talents:
        authors_path = conference_dir / "authors.json"
        if authors_path.exists():
            try:
                authors_data = load_json_file(str(authors_path))
                authors = authors_data.get("authors", [])

                for author in authors:
                    match = False
                    if name and author.get("name", "").lower() == name.lower():
                        match = True
                    if aminer_id and author.get("aminer_id") == aminer_id:
                        match = True
                    if match:
                        # Convert author format to talent format
                        talent = {
                            "name": author.get("name"),
                            "aminer_id": author.get("aminer_id"),
                            "affiliation": author.get("organization"),
                            "roles": [],  # Authors don't have roles
                            "description": None,
                        }
                        matching_talents.append(talent)
            except Exception as e:
                print(f"Error loading authors data: {e}")

    if not matching_talents:
        raise HTTPException(status_code=404, detail="Scholar not found")

    # Build detailed response
    results = []
    for talent in matching_talents:
        scholar_aminer_id = talent.get("aminer_id")
        detail = build_scholar_detail(talent, scholar_aminer_id, conference_id)
        results.append(detail)

    return results


def build_scholar_detail(talent: dict, aminer_id: Optional[str], conference_id: str) -> ScholarDetail:
    """Build detailed scholar information from multiple sources."""
    photo_url = get_scholar_photo(aminer_id)

    # Start with basic info
    detail = ScholarDetail(
        name=talent.get("name", "Unknown"),
        aliases=talent.get("aliases"),
        affiliation=talent.get("affiliation"),
        roles=talent.get("roles", []),
        description=talent.get("description"),
        sources=talent.get("sources"),
        source_urls=talent.get("source_urls"),
        aminer_id=aminer_id,
        photo_url=photo_url,
    )

    # Add aminer_validation if present
    if talent.get("aminer_validation"):
        detail.aminer_validation = AminerValidation(**talent["aminer_validation"])

    # Load AMiner data if available
    if aminer_id:
        aminer_path = settings.aminer_scholars_dir / f"{aminer_id}.json"
        if aminer_path.exists():
            try:
                aminer_data = load_json_file(str(aminer_path))
                aminer_detail = aminer_data.get("detail", {})

                # Name: always use name_zh from AMiner
                detail.name_zh = aminer_detail.get("name_zh")

                # For other fields: prefer Chinese version (_zh), fallback to English
                detail.bio = aminer_detail.get("bio_zh") or aminer_detail.get("bio")
                detail.education = aminer_detail.get("edu_zh") or aminer_detail.get("edu")
                detail.position = aminer_detail.get("position_zh") or aminer_detail.get("position")
                detail.organizations = aminer_detail.get("orgs_zh") or aminer_detail.get("orgs")
                detail.honors = aminer_detail.get("honor")

                # Research interests from figure
                figure = aminer_data.get("figure", {})
                detail.research_interests = figure.get("ai_interests")
            except Exception as e:
                print(f"Error loading AMiner data for {aminer_id}: {e}")

        # Load enriched data if available
        enriched_path = settings.enriched_scholars_dir / f"{aminer_id}.json"
        if enriched_path.exists():
            try:
                enriched_data = load_json_file(str(enriched_path))
                detail.homepage = enriched_data.get("homepage")
                detail.google_scholar = enriched_data.get("google_scholar")
                detail.dblp = enriched_data.get("dblp")
                detail.linkedin = enriched_data.get("linkedin")
                detail.twitter = enriched_data.get("twitter")
                detail.email = enriched_data.get("email")
                detail.orcid = enriched_data.get("orcid")
                detail.semantic_scholar = enriched_data.get("semantic_scholar")
                detail.additional_info = enriched_data.get("additional_info")

                # Load academic indices
                if enriched_data.get("indices"):
                    indices_data = enriched_data["indices"]
                    detail.indices = AcademicIndices(
                        hindex=indices_data.get("hindex"),
                        gindex=indices_data.get("gindex"),
                        citations=indices_data.get("citations"),
                        pubs=indices_data.get("pubs"),
                        activity=indices_data.get("activity"),
                        diversity=indices_data.get("diversity"),
                        sociability=indices_data.get("sociability"),
                    )

                # Load labels if available
                if enriched_data.get("labels"):
                    labels_data = enriched_data["labels"]
                    detail.labels = ScholarLabels(
                        last_updated=labels_data.get("last_updated"),
                        results=[LabelResult(**r) for r in labels_data.get("results", [])]
                    )
            except Exception as e:
                print(f"Error loading enriched data for {aminer_id}: {e}")

    # Load conference papers
    scholar_name_normalized = talent.get("name", "").lower()
    detail.conference_papers = get_scholar_conference_papers(conference_id, scholar_name_normalized)

    return detail


def get_scholar_conference_papers(conference_id: str, scholar_name_normalized: str) -> Optional[list[ConferencePaper]]:
    """Get papers for a scholar in a specific conference."""
    try:
        conference_dir = settings.data_dir / conference_id
        papers_by_author_path = conference_dir / "indexes" / "papers_by_author.json"
        papers_path = conference_dir / "papers.json"
        authors_path = conference_dir / "authors.json"

        print(f"DEBUG: Looking for papers for scholar: '{scholar_name_normalized}'")
        print(f"DEBUG: Conference dir: {conference_dir}")
        print(f"DEBUG: Papers by author path exists: {papers_by_author_path.exists()}")
        print(f"DEBUG: Papers path exists: {papers_path.exists()}")

        if not papers_by_author_path.exists() or not papers_path.exists():
            print(f"DEBUG: Missing required files, returning None")
            return None

        # Load papers by author index
        papers_by_author = load_json_file(str(papers_by_author_path))
        paper_ids = papers_by_author.get(scholar_name_normalized, [])

        print(f"DEBUG: Found paper IDs: {paper_ids}")

        if not paper_ids:
            print(f"DEBUG: No paper IDs found, returning None")
            return None

        # Load all papers
        papers_data = load_json_file(str(papers_path))
        all_papers = papers_data.get("papers", [])

        # Build paper ID to paper mapping
        papers_map = {p["paper_id"]: p for p in all_papers}

        # Load authors data for coauthor lookup
        authors_map = {}
        if authors_path.exists():
            authors_data = load_json_file(str(authors_path))
            for author in authors_data.get("authors", []):
                normalized = author.get("normalized_name", "").lower()
                if normalized:
                    authors_map[normalized] = author

        # Build conference papers list
        conference_papers = []
        for paper_id in paper_ids:
            paper = papers_map.get(paper_id)
            if not paper:
                continue

            # Get abstract from AMiner if available
            abstract = None
            aminer_paper_id = paper.get("aminer_paper_id")
            if aminer_paper_id:
                aminer_paper_path = settings.aminer_papers_dir / f"{aminer_paper_id}.json"
                if aminer_paper_path.exists():
                    try:
                        aminer_paper_data = load_json_file(str(aminer_paper_path))
                        abstract = aminer_paper_data.get("detail", {}).get("abstract")
                    except Exception as e:
                        print(f"Error loading AMiner paper {aminer_paper_id}: {e}")

            # Build authors and coauthors lists, track scholar position
            authors = []
            coauthors = []
            author_position = None
            for index, author_name in enumerate(paper.get("authors", []), start=1):
                author_name_normalized = author_name.lower()
                author_info = authors_map.get(author_name_normalized, {})

                author = ConferencePaperAuthor(
                    name=author_name,
                    aminer_id=author_info.get("aminer_id"),
                    in_conference=bool(author_info.get("aminer_id")),
                )

                # Add to full authors list
                authors.append(author)

                # Check if this is the current scholar
                if author_name_normalized == scholar_name_normalized:
                    author_position = index
                else:
                    # Add to coauthors only if not the current scholar
                    coauthors.append(author)

            # Determine presentation type from source file
            presentation_type = None
            source_file = paper.get("_source_file", "")
            if "oral" in source_file.lower():
                presentation_type = "oral"
            elif "poster" in source_file.lower():
                presentation_type = "poster"

            conference_papers.append(ConferencePaper(
                paper_id=paper["paper_id"],
                title=paper.get("title", ""),
                track=paper.get("track"),
                session=paper.get("session"),
                room=paper.get("room"),
                date=paper.get("date"),
                presentation_type=presentation_type,
                authors=authors,
                coauthors=coauthors,
                author_position=author_position,
                abstract=abstract,
            ))

        return conference_papers if conference_papers else None

    except Exception as e:
        print(f"Error loading conference papers for {scholar_name_normalized}: {e}")
        return None


@app.get("/api/avatar/local/{aminer_id}")
async def get_local_avatar_endpoint(aminer_id: str):
    """
    Get avatar image from local storage (data/aminer/avatars).
    This endpoint serves avatars that were downloaded by the download script.
    """
    # Check if we have a local avatar
    local_avatar_path = get_local_avatar_path(aminer_id)
    if not local_avatar_path:
        raise HTTPException(status_code=404, detail="Avatar not found locally")

    # Determine content type from extension
    ext = local_avatar_path.suffix.lower()
    content_types = {
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
        ".gif": "image/gif",
        ".webp": "image/webp",
    }
    content_type = content_types.get(ext, "image/jpeg")

    return FileResponse(local_avatar_path, media_type=content_type)


@app.get("/api/avatar/{aminer_id}")
async def get_avatar(aminer_id: str):
    """
    Get avatar image for a scholar.
    Caches the image locally to avoid repeated remote requests.
    Also caches failed fetches to avoid slow repeated timeouts.
    """
    # Get the original photo URL
    photo_url = get_scholar_photo_url(aminer_id)
    if not photo_url:
        raise HTTPException(status_code=404, detail="No photo URL found")

    # Check if we have a cached version
    cache_path = get_avatar_cache_path(photo_url)
    if cache_path.exists():
        # Determine content type from extension
        ext = cache_path.suffix.lower()
        content_types = {
            ".jpg": "image/jpeg",
            ".jpeg": "image/jpeg",
            ".png": "image/png",
            ".gif": "image/gif",
            ".webp": "image/webp",
        }
        content_type = content_types.get(ext, "image/jpeg")
        return FileResponse(cache_path, media_type=content_type)

    # Check if this URL has failed recently
    if is_avatar_fetch_failed(photo_url):
        raise HTTPException(status_code=404, detail="Avatar fetch previously failed")

    # Fetch from remote and cache
    try:
        response = await http_client.get(photo_url)
        if response.status_code != 200:
            mark_avatar_fetch_failed(photo_url)
            raise HTTPException(status_code=404, detail="Remote avatar not found")

        # Get content type
        content_type = response.headers.get("content-type", "image/jpeg")

        # Save to cache
        with open(cache_path, "wb") as f:
            f.write(response.content)

        return Response(content=response.content, media_type=content_type)

    except httpx.RequestError as e:
        mark_avatar_fetch_failed(photo_url)
        raise HTTPException(status_code=502, detail=f"Failed to fetch avatar: {str(e)}")


@app.get("/api/labels", response_model=LabelsConfig)
def get_labels_config():
    """
    Get labels configuration.
    Returns all available label definitions from config/labels.json.
    """
    if not settings.labels_config_path.exists():
        raise HTTPException(status_code=404, detail="Labels configuration not found")

    try:
        labels_data = load_json_file(str(settings.labels_config_path))
        return LabelsConfig(**labels_data)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error loading labels config: {e}")


@app.get("/api/conferences/{conference_id}/people/filter")
def filter_people_by_labels(
    conference_id: str,
    labels: Optional[str] = Query(None, description="Label filters in format 'name:value,name:value' (e.g., 'Chinese:true,Student:false')"),
):
    """
    Filter people (authors + scholars) by label values.
    Returns merged data from both authors and scholars where labels match with high confidence.
    """
    conference_dir = settings.data_dir / conference_id

    if not conference_dir.exists():
        raise HTTPException(status_code=404, detail=f"Conference not found: {conference_id}")

    # Parse label filters
    label_filters: dict[str, bool] = {}
    if labels:
        for filter_item in labels.split(","):
            if ":" in filter_item:
                name, value = filter_item.split(":", 1)
                name = name.strip()
                value = value.strip().lower()
                if value in ("true", "false"):
                    label_filters[name] = value == "true"

    # Load scholars data
    scholars_path = conference_dir / "scholars.json"
    talents = []
    if scholars_path.exists():
        try:
            scholars_data = load_json_file(str(scholars_path))
            talents = scholars_data.get("talents", [])
        except Exception as e:
            logger.error(f"Error loading scholars data: {e}")

    # Load authors data
    authors_path = conference_dir / "authors.json"
    authors = []
    if authors_path.exists():
        try:
            authors_data = load_json_file(str(authors_path))
            authors = authors_data.get("authors", [])
        except Exception as e:
            logger.error(f"Error loading authors data: {e}")

    # Create a set of all unique aminer_ids from both sources
    all_aminer_ids = set()
    for talent in talents:
        if talent.get("aminer_id"):
            all_aminer_ids.add(talent.get("aminer_id"))
    for author in authors:
        if author.get("aminer_id"):
            all_aminer_ids.add(author.get("aminer_id"))

    # Filter by labels if filters are provided
    filtered_aminer_ids = set()
    if label_filters:
        for aminer_id in all_aminer_ids:
            # Load enriched data to check labels
            enriched_path = settings.enriched_scholars_dir / f"{aminer_id}.json"
            if not enriched_path.exists():
                continue

            try:
                enriched_data = load_json_file(str(enriched_path))
                labels_data = enriched_data.get("labels", {})
                results = labels_data.get("results", [])

                # Check if all filter conditions are met with medium or high confidence
                all_match = True
                for label_name, expected_value in label_filters.items():
                    found_match = False
                    for result in results:
                        if result.get("name") == label_name:
                            if (result.get("value") == expected_value and
                                result.get("confidence") in ("high", "medium")):
                                found_match = True
                            break
                    if not found_match:
                        all_match = False
                        break

                if all_match:
                    filtered_aminer_ids.add(aminer_id)
            except Exception:
                continue
    else:
        # No filters, include all
        filtered_aminer_ids = all_aminer_ids

    # Build merged result with filtered aminer_ids
    people_map = {}

    # Add scholars first
    for talent in talents:
        aminer_id = talent.get("aminer_id")
        if aminer_id and aminer_id in filtered_aminer_ids:
            photo_url = get_scholar_photo(aminer_id)

            # Try to get Chinese name from AMiner data
            name_zh = None
            aminer_path = settings.aminer_scholars_dir / f"{aminer_id}.json"
            if aminer_path.exists():
                try:
                    aminer_data = load_json_file(str(aminer_path))
                    name_zh = aminer_data.get("detail", {}).get("name_zh")
                except Exception:
                    pass

            people_map[aminer_id] = {
                "name": talent.get("name", "Unknown"),
                "name_zh": name_zh,
                "affiliation": talent.get("affiliation"),
                "roles": talent.get("roles", []),
                "aminer_id": aminer_id,
                "photo_url": photo_url,
                "description": talent.get("description"),
            }

    # Merge with authors (add metrics)
    for author in authors:
        aminer_id = author.get("aminer_id")
        if aminer_id and aminer_id in filtered_aminer_ids:
            if aminer_id in people_map:
                # Merge data
                existing = people_map[aminer_id]
                people_map[aminer_id] = {
                    **existing,
                    "paper_count": author.get("paper_count"),
                    "h_index": author.get("h_index"),
                    "n_citation": author.get("n_citation"),
                    "n_pubs": author.get("n_pubs"),
                    "organization": author.get("organization"),
                    "organization_zh": author.get("organization_zh"),
                    "name_zh": existing.get("name_zh") or author.get("name_zh"),
                }
            else:
                # Add new author
                photo_url = get_scholar_photo(aminer_id)

                # Try to get Chinese name from AMiner data
                name_zh = author.get("name_zh")
                if not name_zh and aminer_id:
                    aminer_path = settings.aminer_scholars_dir / f"{aminer_id}.json"
                    if aminer_path.exists():
                        try:
                            aminer_data = load_json_file(str(aminer_path))
                            name_zh = aminer_data.get("detail", {}).get("name_zh")
                        except Exception:
                            pass

                people_map[aminer_id] = {
                    "name": author.get("name", "Unknown"),
                    "name_zh": name_zh,
                    "affiliation": author.get("organization"),
                    "roles": [],
                    "aminer_id": aminer_id,
                    "photo_url": photo_url,
                    "description": None,
                    "paper_count": author.get("paper_count"),
                    "h_index": author.get("h_index"),
                    "n_citation": author.get("n_citation"),
                    "n_pubs": author.get("n_pubs"),
                    "organization": author.get("organization"),
                    "organization_zh": author.get("organization_zh"),
                }

    return list(people_map.values())


@app.get("/api/conferences/{conference_id}/scholars/filter", response_model=list[ScholarBasic])
def filter_scholars_by_labels(
    conference_id: str,
    labels: Optional[str] = Query(None, description="Label filters in format 'name:value,name:value' (e.g., 'Chinese:true,Student:false')"),
):
    """
    Filter scholars by label values.
    Only returns scholars where labels match with high confidence.
    """
    conference_dir = settings.data_dir / conference_id
    scholars_path = conference_dir / "scholars.json"

    if not conference_dir.exists():
        raise HTTPException(status_code=404, detail=f"Conference not found: {conference_id}")

    if not scholars_path.exists():
        raise HTTPException(status_code=404, detail=f"Scholars data not found for conference: {conference_id}")

    try:
        scholars_data = load_json_file(str(scholars_path))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error loading scholars data: {e}")

    talents = scholars_data.get("talents", [])

    # Parse label filters
    label_filters: dict[str, bool] = {}
    if labels:
        for filter_item in labels.split(","):
            if ":" in filter_item:
                name, value = filter_item.split(":", 1)
                name = name.strip()
                value = value.strip().lower()
                if value in ("true", "false"):
                    label_filters[name] = value == "true"

    # If no filters, return all scholars
    if not label_filters:
        scholars = []
        for talent in talents:
            aminer_id = talent.get("aminer_id")
            photo_url = get_scholar_photo(aminer_id)

            # Try to get Chinese name from AMiner data
            name_zh = None
            if aminer_id:
                aminer_path = settings.aminer_scholars_dir / f"{aminer_id}.json"
                if aminer_path.exists():
                    try:
                        aminer_data = load_json_file(str(aminer_path))
                        name_zh = aminer_data.get("detail", {}).get("name_zh")
                    except Exception:
                        pass

            scholars.append(ScholarBasic(
                name=talent.get("name", "Unknown"),
                name_zh=name_zh,
                affiliation=talent.get("affiliation"),
                roles=talent.get("roles", []),
                aminer_id=aminer_id,
                photo_url=photo_url,
                description=talent.get("description"),
            ))
        return scholars

    # Filter scholars by labels
    filtered_scholars = []
    for talent in talents:
        aminer_id = talent.get("aminer_id")
        if not aminer_id:
            continue

        # Load enriched data to check labels
        enriched_path = settings.enriched_scholars_dir / f"{aminer_id}.json"
        if not enriched_path.exists():
            continue

        try:
            enriched_data = load_json_file(str(enriched_path))
            labels_data = enriched_data.get("labels", {})
            results = labels_data.get("results", [])

            # Check if all filter conditions are met with medium or high confidence
            all_match = True
            for label_name, expected_value in label_filters.items():
                found_match = False
                for result in results:
                    if result.get("name") == label_name:
                        if (result.get("value") == expected_value and
                            result.get("confidence") in ("high", "medium")):
                            found_match = True
                        break
                if not found_match:
                    all_match = False
                    break

            if all_match:
                photo_url = get_scholar_photo(aminer_id)

                # Try to get Chinese name from AMiner data
                name_zh = None
                aminer_path = settings.aminer_scholars_dir / f"{aminer_id}.json"
                if aminer_path.exists():
                    try:
                        aminer_data = load_json_file(str(aminer_path))
                        name_zh = aminer_data.get("detail", {}).get("name_zh")
                    except Exception:
                        pass

                filtered_scholars.append(ScholarBasic(
                    name=talent.get("name", "Unknown"),
                    name_zh=name_zh,
                    affiliation=talent.get("affiliation"),
                    roles=talent.get("roles", []),
                    aminer_id=aminer_id,
                    photo_url=photo_url,
                    description=talent.get("description"),
                ))
        except Exception:
            continue

    return filtered_scholars


def clean_excel_value(value):
    """
    Clean value for Excel export.
    Remove characters that are not allowed in Excel cells.
    """
    if value is None:
        return None
    if not isinstance(value, str):
        return value

    # Remove control characters (except tab, newline, carriage return)
    # Excel doesn't allow characters with ASCII codes 0-31 except 9, 10, 13
    cleaned = ""
    for char in value:
        code = ord(char)
        if code >= 32 or code in (9, 10, 13):
            cleaned += char

    # Replace HTML line breaks with newlines
    cleaned = cleaned.replace("<br><br>", "\n\n")
    cleaned = cleaned.replace("<br>", "\n")

    return cleaned


def collect_person_data(person: dict, conference_id: str, role: str = None) -> dict:
    """
    Collect comprehensive data for a person (scholar or author).

    Args:
        person: Basic person data from scholars.json or authors.json
        conference_id: Conference ID
        role: Role string (for scholars) or None (will use "Paper Author" for authors)

    Returns:
        Dictionary with all person data for Excel export
    """
    aminer_id = person.get("aminer_id")
    data = {
        "name": person.get("name", ""),
        "name_zh": None,
        "role": role if role else "Paper Author",
        "first_org_en": None,
        "first_org_zh": None,
        "all_orgs_en": None,
        "all_orgs_zh": None,
        "bio_en": None,
        "bio_zh": None,
        "position_en": None,
        "position_zh": None,
        "education_en": None,
        "education_zh": None,
        "homepage": None,
        "google_scholar": None,
        "aminer_id": aminer_id or "",
        "dblp": None,
        "website_url": f"https://scholar.ihainan.me/conference/{conference_id}/scholar?aminer_id={aminer_id}&from=people" if aminer_id else "",
        "hindex": None,
        "gindex": None,
        "citations": None,
        "pubs": None,
        "activity": None,
        "diversity": None,
        "sociability": None,
        "first_email": None,
        "all_emails": None,
        "research_tags": None,
        "labels": {}
    }

    # Load AMiner data
    if aminer_id:
        aminer_path = settings.aminer_scholars_dir / f"{aminer_id}.json"
        if aminer_path.exists():
            try:
                aminer_data = load_json_file(str(aminer_path))
                detail = aminer_data.get("detail", {})

                data["name_zh"] = detail.get("name_zh")
                data["bio_en"] = detail.get("bio")
                data["bio_zh"] = detail.get("bio_zh")
                data["position_en"] = detail.get("position")
                data["position_zh"] = detail.get("position_zh")
                data["education_en"] = detail.get("edu")
                data["education_zh"] = detail.get("edu_zh")

                # Parse organizations
                orgs = detail.get("orgs", [])
                org_zhs = detail.get("org_zhs", [])
                if orgs and len(orgs) > 0:
                    all_orgs = orgs[0]
                    data["all_orgs_en"] = all_orgs
                    # Split by semicolon to get first org
                    if ";" in all_orgs:
                        data["first_org_en"] = all_orgs.split(";")[0].strip()
                    else:
                        data["first_org_en"] = all_orgs

                if org_zhs and len(org_zhs) > 0:
                    all_orgs_zh = org_zhs[0]
                    data["all_orgs_zh"] = all_orgs_zh
                    if ";" in all_orgs_zh:
                        data["first_org_zh"] = all_orgs_zh.split(";")[0].strip()
                    else:
                        data["first_org_zh"] = all_orgs_zh
            except Exception as e:
                logger.error(f"Error loading AMiner data for {aminer_id}: {e}")

        # Load enriched data
        enriched_path = settings.enriched_scholars_dir / f"{aminer_id}.json"
        if enriched_path.exists():
            try:
                enriched_data = load_json_file(str(enriched_path))

                data["homepage"] = enriched_data.get("homepage")
                data["google_scholar"] = enriched_data.get("google_scholar")
                data["dblp"] = enriched_data.get("dblp")

                # Parse emails
                email = enriched_data.get("email")
                if email:
                    data["all_emails"] = email
                    if ";" in email:
                        data["first_email"] = email.split(";")[0].strip()
                    else:
                        data["first_email"] = email

                # Academic indices
                indices = enriched_data.get("indices", {})
                data["hindex"] = indices.get("hindex")
                data["gindex"] = indices.get("gindex")
                data["citations"] = indices.get("citations")
                data["pubs"] = indices.get("pubs")
                data["activity"] = indices.get("activity")
                data["diversity"] = indices.get("diversity")
                data["sociability"] = indices.get("sociability")

                # Research tags
                research_tags = enriched_data.get("research_tags", [])
                if research_tags:
                    data["research_tags"] = ", ".join(research_tags)

                # Labels
                labels_data = enriched_data.get("labels", {})
                results = labels_data.get("results", [])
                for result in results:
                    label_name = result.get("name")
                    if label_name:
                        data["labels"][label_name] = {
                            "value": result.get("value"),
                            "confidence": result.get("confidence"),
                            "reason": result.get("reason")
                        }
            except Exception as e:
                logger.error(f"Error loading enriched data for {aminer_id}: {e}")

    return data


def create_excel_export(conference_id: str) -> BytesIO:
    """
    Create Excel file with conference scholars and authors data.

    Args:
        conference_id: Conference ID

    Returns:
        BytesIO object containing the Excel file
    """
    conference_dir = settings.data_dir / conference_id

    if not conference_dir.exists():
        raise HTTPException(status_code=404, detail=f"Conference not found: {conference_id}")

    # Load label definitions
    labels_config = []
    if settings.labels_config_path.exists():
        try:
            labels_data = load_json_file(str(settings.labels_config_path))
            labels_config = labels_data.get("labels", [])
        except Exception as e:
            logger.error(f"Error loading labels config: {e}")

    # Define base columns (name, width)
    base_columns = [
        ("编号", 8),
        ("姓名（英文）", 20),
        ("姓名（中文）", 15),
        ("会议角色", 25),
        ("所属机构（主要，英文）", 40),
        ("所属机构（主要，中文）", 40),
        ("所属机构（全部，英文）", 60),
        ("所属机构（全部，中文）", 60),
        ("个人简介（英文）", 80),
        ("个人简介（中文）", 80),
        ("职位/职称（英文）", 25),
        ("职位/职称（中文）", 25),
        ("教育经历（英文）", 80),
        ("教育经历（中文）", 80),
        ("个人主页", 50),
        ("谷歌学术主页", 50),
        ("AMiner ID", 30),
        ("DBLP主页", 50),
        ("网站个人主页", 80),
        ("H指数", 12),
        ("G指数", 12),
        ("引用数", 12),
        ("发表论文数", 10),
        ("活跃度", 12),
        ("多样性", 12),
        ("社交度", 12),
        ("邮箱（主要）", 35),
        ("邮箱（全部）", 60),
        ("研究方向", 80),
    ]

    # Add label columns
    label_columns = []
    for label_def in labels_config:
        label_name = label_def.get("name", "")
        label_columns.extend([
            (f"{label_name}_value", 15),
            (f"{label_name}_confidence", 15),
            (f"{label_name}_reason", 80),
        ])

    all_columns = base_columns + label_columns

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Define fonts
    header_font = Font(name='等线', size=12, bold=True)
    data_font = Font(name='等线', size=12)

    # Process authors data first (will be first sheet)
    authors_path = conference_dir / "authors.json"
    authors_sheet_data = []
    if authors_path.exists():
        try:
            authors_data = load_json_file(str(authors_path))
            authors = authors_data.get("authors", [])

            # Collect all author data
            for author in authors:
                person_data = collect_person_data(author, conference_id, "Paper Author")
                authors_sheet_data.append(person_data)
        except Exception as e:
            logger.error(f"Error loading authors data: {e}")

    # Sort authors by citations (descending)
    authors_sheet_data.sort(key=lambda x: x.get("citations") or 0, reverse=True)

    # Create Paper Authors sheet (first sheet)
    if authors_sheet_data:
        ws_authors = wb.create_sheet("Paper Authors", 0)

        # Write header
        for col_idx, (col_name, col_width) in enumerate(all_columns, start=1):
            cell = ws_authors.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws_authors.column_dimensions[get_column_letter(col_idx)].width = col_width

        ws_authors.freeze_panes = "A2"  # Freeze header row

        # Add auto filter to all columns
        ws_authors.auto_filter.ref = f"A1:{get_column_letter(len(all_columns))}1"

        # Write data
        for row_idx, person_data in enumerate(authors_sheet_data, start=2):
            row_data = [
                row_idx - 1,  # 编号
                person_data["name"],
                person_data["name_zh"],
                person_data["role"],
                person_data["first_org_en"],
                person_data["first_org_zh"],
                person_data["all_orgs_en"],
                person_data["all_orgs_zh"],
                person_data["bio_en"],
                person_data["bio_zh"],
                person_data["position_en"],
                person_data["position_zh"],
                person_data["education_en"],
                person_data["education_zh"],
                person_data["homepage"],
                person_data["google_scholar"],
                person_data["aminer_id"],
                person_data["dblp"],
                person_data["website_url"],
                person_data["hindex"],
                person_data["gindex"],
                person_data["citations"],
                person_data["pubs"],
                person_data["activity"],
                person_data["diversity"],
                person_data["sociability"],
                person_data["first_email"],
                person_data["all_emails"],
                person_data["research_tags"],
            ]

            # Add label data
            for label_def in labels_config:
                label_name = label_def.get("name", "")
                label_data = person_data["labels"].get(label_name, {})
                row_data.extend([
                    label_data.get("value"),
                    label_data.get("confidence"),
                    label_data.get("reason"),
                ])

            for col_idx, value in enumerate(row_data, start=1):
                cleaned_value = clean_excel_value(value)
                cell = ws_authors.cell(row=row_idx, column=col_idx, value=cleaned_value)
                cell.font = data_font
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Process scholars data
    scholars_path = conference_dir / "scholars.json"
    if scholars_path.exists():
        try:
            scholars_data = load_json_file(str(scholars_path))
            talents = scholars_data.get("talents", [])

            # Collect all scholar data
            scholars_sheet_data = []
            for talent in talents:
                roles = talent.get("roles", [])
                role = roles[0] if roles else ""
                person_data = collect_person_data(talent, conference_id, role)
                scholars_sheet_data.append(person_data)

            # Sort scholars by citations (descending)
            scholars_sheet_data.sort(key=lambda x: x.get("citations") or 0, reverse=True)

            ws_scholars = wb.create_sheet("Conference Organizers")

            # Write header
            for col_idx, (col_name, col_width) in enumerate(all_columns, start=1):
                cell = ws_scholars.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws_scholars.column_dimensions[get_column_letter(col_idx)].width = col_width

            ws_scholars.freeze_panes = "A2"  # Freeze header row

            # Add auto filter to all columns
            ws_scholars.auto_filter.ref = f"A1:{get_column_letter(len(all_columns))}1"

            # Write data
            for row_idx, person_data in enumerate(scholars_sheet_data, start=2):
                row_data = [
                    row_idx - 1,  # 编号
                    person_data["name"],
                    person_data["name_zh"],
                    person_data["role"],
                    person_data["first_org_en"],
                    person_data["first_org_zh"],
                    person_data["all_orgs_en"],
                    person_data["all_orgs_zh"],
                    person_data["bio_en"],
                    person_data["bio_zh"],
                    person_data["position_en"],
                    person_data["position_zh"],
                    person_data["education_en"],
                    person_data["education_zh"],
                    person_data["homepage"],
                    person_data["google_scholar"],
                    person_data["aminer_id"],
                    person_data["dblp"],
                    person_data["website_url"],
                    person_data["hindex"],
                    person_data["gindex"],
                    person_data["citations"],
                    person_data["pubs"],
                    person_data["activity"],
                    person_data["diversity"],
                    person_data["sociability"],
                    person_data["first_email"],
                    person_data["all_emails"],
                    person_data["research_tags"],
                ]

                # Add label data
                for label_def in labels_config:
                    label_name = label_def.get("name", "")
                    label_data = person_data["labels"].get(label_name, {})
                    row_data.extend([
                        label_data.get("value"),
                        label_data.get("confidence"),
                        label_data.get("reason"),
                    ])

                for col_idx, value in enumerate(row_data, start=1):
                    cleaned_value = clean_excel_value(value)
                    cell = ws_scholars.cell(row=row_idx, column=col_idx, value=cleaned_value)
                    cell.font = data_font
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
        except Exception as e:
            logger.error(f"Error processing scholars data: {e}")

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@app.get("/api/conferences/{conference_id}/export/excel")
def export_conference_excel(conference_id: str):
    """
    Export conference scholars and authors data to Excel.
    Returns an Excel file with two sheets: Conference Organizers and Paper Authors.
    """
    try:
        excel_file = create_excel_export(conference_id)

        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={conference_id}_export.xlsx"
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating Excel export: {e}")
        raise HTTPException(status_code=500, detail=f"Error creating Excel export: {str(e)}")


@app.post("/api/cache/clear")
def clear_cache():
    """Clear all cached data (JSON cache only, not avatar cache)."""
    _load_json_file.cache_clear()
    _cache_timestamp.clear()
    return {"status": "cache cleared"}


@app.post("/api/avatar/cache/clear")
def clear_avatar_cache():
    """Clear all cached avatar images."""
    count = 0
    for file in settings.avatar_cache_dir.iterdir():
        if file.is_file():
            file.unlink()
            count += 1
    return {"status": "avatar cache cleared", "files_deleted": count}


# Lifecycle events
@app.on_event("startup")
async def startup_event():
    """Application startup tasks."""
    logger.info("Starting Conference API Service...")
    logger.info(f"Configuration:")
    logger.info(f"  - Host: {settings.host}:{settings.port}")
    logger.info(f"  - Data Dir: {settings.data_dir}")
    logger.info(f"  - Config Dir: {settings.config_dir}")
    logger.info(f"  - Avatar Cache Dir: {settings.avatar_cache_dir}")
    logger.info(f"  - CORS Origins: {settings.cors_origins}")
    logger.info(f"  - Log Level: {settings.log_level}")
    logger.info("Service started successfully")


@app.on_event("shutdown")
async def shutdown_event():
    """Application shutdown tasks."""
    logger.info("Shutting down Conference API Service...")
    await http_client.aclose()
    logger.info("Service stopped")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(
        "main:app",
        host=settings.host,
        port=settings.port,
        reload=True,
        reload_excludes=["avatar_cache/*"]
    )
